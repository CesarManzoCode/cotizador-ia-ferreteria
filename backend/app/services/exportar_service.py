"""
Servicio de exportación de cotizaciones a Excel.
Genera un archivo Excel formateado y profesional listo para entregar al cliente.

PUNTOS DE DEBUGGING:
- Si los colores/estilos no se ven bien, ajusta las constantes de estilo abajo
- Si necesitas agregar columnas, modifica _escribir_filas_productos()
- Si necesitas cambiar el orden de columnas, modifica COLUMNAS_COTIZACION

CÓMO MODIFICAR EL FORMATO DEL EXCEL DE SALIDA:
1. Colores: modifica las constantes COLOR_* abajo
2. Columnas: modifica COLUMNAS_COTIZACION
3. Ancho de columnas: modifica ANCHOS_COLUMNA
4. Agregar logo: busca el comentario "# FUTURA: logo de empresa"
"""

import logging
import os
import uuid
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.schemas.cotizacion import NivelConfianza, ProductoEncontrado

logger = logging.getLogger(__name__)

# ── Constantes de estilo ───────────────────────────────────────────────────────
# Modifica aquí para cambiar la apariencia del Excel de salida

COLOR_ENCABEZADO_FONDO = "1E293B"      # Azul oscuro (fondo de headers)
COLOR_ENCABEZADO_TEXTO = "FFFFFF"      # Blanco (texto de headers)
COLOR_FILA_PAR = "F8FAFC"             # Gris muy claro (filas pares)
COLOR_FILA_IMPAR = "FFFFFF"           # Blanco (filas impares)
COLOR_DUDOSO_FONDO = "FEF9C3"        # Amarillo suave (advertencia dudosa)
COLOR_NO_ENCONTRADO_FONDO = "FEE2E2" # Rojo suave (no encontrado)
COLOR_TOTAL_FONDO = "1E293B"         # Azul oscuro (fila total)
COLOR_TOTAL_TEXTO = "FFFFFF"         # Blanco (texto total)
COLOR_TITULO_TEXTO = "1E293B"        # Azul oscuro (título principal)

# ── Definición de columnas del Excel de cotización ────────────────────────────
# Modifica aquí para agregar/quitar/reordenar columnas
COLUMNAS_COTIZACION = [
    ("Producto Solicitado", 30),
    ("Producto Encontrado", 35),
    ("Marca", 15),
    ("Unidad", 10),
    ("Cantidad", 10),
    ("Precio Unitario", 15),
    ("Subtotal", 15),
    ("Código", 15),
    ("Observaciones", 40),
]


def _obtener_borde_delgado():
    """Crea un estilo de borde delgado para celdas."""
    lado = Side(style="thin", color="E2E8F0")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _crear_dir_temporal():
    """Crea el directorio temporal para Excel generados si no existe."""
    dir_temp = Path(settings.EXPORTAR_DIR_TEMPORAL)
    dir_temp.mkdir(parents=True, exist_ok=True)
    return dir_temp


class ExportadorExcel:
    """
    Genera el archivo Excel de cotización formateado.
    """

    def __init__(self):
        self.dir_temporal = _crear_dir_temporal()

    def generar_excel_cotizacion(
        self,
        productos: List[ProductoEncontrado],
        texto_solicitud: str,
        cotizacion_id: Optional[int] = None,
    ) -> str:
        """
        Genera el archivo Excel de cotización y lo guarda en disco.
        
        Returns:
            Nombre del archivo generado (no la ruta completa)
        """
        # Nombre único para el archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        uid_corto = str(uuid.uuid4())[:8]
        nombre_archivo = f"{settings.EXPORTAR_NOMBRE_BASE}_{timestamp}_{uid_corto}.xlsx"
        ruta_completa = self.dir_temporal / nombre_archivo

        wb = Workbook()
        ws = wb.active
        ws.title = "Cotización"

        fila_actual = 1

        # ── Encabezado del documento ───────────────────────────────────────────
        fila_actual = self._escribir_encabezado_documento(ws, fila_actual, texto_solicitud, cotizacion_id)

        # ── Encabezados de columnas ────────────────────────────────────────────
        fila_actual = self._escribir_encabezados_columnas(ws, fila_actual)

        # ── Filas de productos ─────────────────────────────────────────────────
        fila_inicio_datos = fila_actual
        fila_actual = self._escribir_filas_productos(ws, fila_actual, productos)

        # ── Totales ────────────────────────────────────────────────────────────
        fila_actual = self._escribir_totales(ws, fila_actual, productos)

        # ── Notas y leyenda ────────────────────────────────────────────────────
        fila_actual = self._escribir_notas(ws, fila_actual + 1, productos)

        # ── Ajustar anchos de columnas ─────────────────────────────────────────
        self._ajustar_anchos(ws)

        # Guardar
        wb.save(ruta_completa)
        logger.info(f"Excel de cotización generado: {ruta_completa}")

        return nombre_archivo

    def _escribir_encabezado_documento(self, ws, fila: int, texto_solicitud: str, cotizacion_id) -> int:
        """Escribe el bloque de encabezado con datos de la empresa y cotización."""

        # FUTURA: logo de empresa
        # from openpyxl.drawing.image import Image
        # logo = Image("logo.png"); ws.add_image(logo, "A1")

        # Título principal
        ws.merge_cells(f"A{fila}:I{fila}")
        celda_titulo = ws[f"A{fila}"]
        celda_titulo.value = "COTIZACIÓN"
        celda_titulo.font = Font(name="Calibri", size=20, bold=True, color=COLOR_TITULO_TEXTO)
        celda_titulo.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[fila].height = 35
        fila += 1

        # Fecha y número de cotización
        ws.merge_cells(f"A{fila}:E{fila}")
        ws[f"A{fila}"].value = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws[f"A{fila}"].font = Font(name="Calibri", size=10, color="64748B")

        ws.merge_cells(f"F{fila}:I{fila}")
        id_texto = f"No. {cotizacion_id}" if cotizacion_id else "Borrador"
        ws[f"F{fila}"].value = f"Cotización: {id_texto}"
        ws[f"F{fila}"].font = Font(name="Calibri", size=10, bold=True, color="64748B")
        ws[f"F{fila}"].alignment = Alignment(horizontal="right")
        fila += 1

        # Solicitud original
        ws.merge_cells(f"A{fila}:I{fila}")
        ws[f"A{fila}"].value = f"Solicitud: {texto_solicitud[:200]}"
        ws[f"A{fila}"].font = Font(name="Calibri", size=9, italic=True, color="94A3B8")
        ws[f"A{fila}"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[fila].height = 25
        fila += 2  # Línea en blanco antes de columnas

        return fila

    def _escribir_encabezados_columnas(self, ws, fila: int) -> int:
        """Escribe la fila de encabezados de columnas con estilo."""
        fill_header = PatternFill("solid", fgColor=COLOR_ENCABEZADO_FONDO)
        font_header = Font(name="Calibri", size=10, bold=True, color=COLOR_ENCABEZADO_TEXTO)
        alineacion_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, (nombre, ancho) in enumerate(COLUMNAS_COTIZACION, start=1):
            celda = ws.cell(row=fila, column=col_idx, value=nombre)
            celda.fill = fill_header
            celda.font = font_header
            celda.alignment = alineacion_centro
            celda.border = _obtener_borde_delgado()

        ws.row_dimensions[fila].height = 30
        return fila + 1

    def _escribir_filas_productos(self, ws, fila: int, productos: List[ProductoEncontrado]) -> int:
        """Escribe una fila por cada producto en la cotización."""
        borde = _obtener_borde_delgado()

        for i, producto in enumerate(productos):
            es_par = i % 2 == 0

            # Color de fondo según confianza y alternancia
            if producto.nivel_confianza == NivelConfianza.DUDOSO:
                color_fondo = COLOR_DUDOSO_FONDO
            elif producto.nivel_confianza == NivelConfianza.NO_ENCONTRADO:
                color_fondo = COLOR_NO_ENCONTRADO_FONDO
            else:
                color_fondo = COLOR_FILA_PAR if es_par else COLOR_FILA_IMPAR

            fill = PatternFill("solid", fgColor=color_fondo)
            font_normal = Font(name="Calibri", size=9)
            font_numero = Font(name="Calibri", size=9)
            alin_centro = Alignment(horizontal="center", vertical="center")
            alin_dcha = Alignment(horizontal="right", vertical="center")
            alin_izq = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # Observaciones
            observacion = ""
            if producto.advertencia:
                observacion = producto.advertencia
            elif producto.nivel_confianza == NivelConfianza.ALTO:
                observacion = f"Coincidencia {producto.score_similitud:.0f}%"

            # Formatear precio y subtotal
            precio_str = f"${producto.precio_unitario:,.2f}" if producto.precio_unitario else "—"
            subtotal_str = f"${producto.subtotal:,.2f}" if producto.subtotal else "—"

            valores = [
                (producto.nombre_solicitado, alin_izq),
                (producto.nombre_encontrado or "Sin coincidencia", alin_izq),
                (producto.marca or "—", alin_centro),
                (producto.unidad or "—", alin_centro),
                (producto.cantidad, alin_centro),
                (precio_str, alin_dcha),
                (subtotal_str, alin_dcha),
                (producto.codigo_ferrol or "—", alin_centro),
                (observacion, alin_izq),
            ]

            for col_idx, (valor, alineacion) in enumerate(valores, start=1):
                celda = ws.cell(row=fila, column=col_idx, value=valor)
                celda.fill = fill
                celda.font = font_normal
                celda.alignment = alineacion
                celda.border = borde

            ws.row_dimensions[fila].height = 20
            fila += 1

        return fila

    def _escribir_totales(self, ws, fila: int, productos: List[ProductoEncontrado]) -> int:
        """Escribe la fila de totales al final de los productos."""
        total = sum(p.subtotal or 0 for p in productos if p.nivel_confianza != NivelConfianza.NO_ENCONTRADO)

        fill_total = PatternFill("solid", fgColor=COLOR_TOTAL_FONDO)
        font_total = Font(name="Calibri", size=10, bold=True, color=COLOR_TOTAL_TEXTO)
        borde = _obtener_borde_delgado()

        # Fila vacía separadora
        fila += 1

        # Fila total
        ws.merge_cells(f"A{fila}:F{fila}")
        celda_etiq = ws[f"A{fila}"]
        celda_etiq.value = "SUBTOTAL COTIZACIÓN (productos encontrados)"
        celda_etiq.fill = fill_total
        celda_etiq.font = font_total
        celda_etiq.alignment = Alignment(horizontal="right", vertical="center")
        celda_etiq.border = borde

        celda_total = ws[f"G{fila}"]
        celda_total.value = f"${total:,.2f}"
        celda_total.fill = fill_total
        celda_total.font = font_total
        celda_total.alignment = Alignment(horizontal="right", vertical="center")
        celda_total.border = borde

        # Rellenar celdas H e I de la fila total
        for col in ["H", "I"]:
            c = ws[f"{col}{fila}"]
            c.fill = fill_total
            c.border = borde

        ws.row_dimensions[fila].height = 22
        return fila + 1

    def _escribir_notas(self, ws, fila: int, productos: List[ProductoEncontrado]) -> int:
        """Escribe la leyenda de colores y notas al pie."""
        font_nota = Font(name="Calibri", size=8, italic=True, color="64748B")

        notas = [
            "Leyenda:",
            "  🟡 Fondo amarillo: coincidencia dudosa — verificar manualmente antes de confirmar",
            "  🔴 Fondo rojo: producto no encontrado en catálogo — requiere cotización manual",
            "  Precios mostrados corresponden a PRECIO PÚBLICO NETO del catálogo",
            "  Esta cotización es un borrador generado automáticamente y requiere revisión",
        ]

        for nota in notas:
            ws.merge_cells(f"A{fila}:I{fila}")
            ws[f"A{fila}"].value = nota
            ws[f"A{fila}"].font = font_nota
            fila += 1

        return fila

    def _ajustar_anchos(self, ws):
        """Aplica los anchos de columna definidos en COLUMNAS_COTIZACION."""
        for col_idx, (_, ancho) in enumerate(COLUMNAS_COTIZACION, start=1):
            letra = get_column_letter(col_idx)
            ws.column_dimensions[letra].width = ancho

    def obtener_ruta_archivo(self, nombre_archivo: str) -> Path:
        """Retorna la ruta completa de un archivo generado."""
        return self.dir_temporal / nombre_archivo

    def archivo_existe(self, nombre_archivo: str) -> bool:
        """Verifica si un archivo generado existe en disco."""
        return (self.dir_temporal / nombre_archivo).exists()


# Importación necesaria para Optional en el método
from typing import Optional  # noqa: E402
