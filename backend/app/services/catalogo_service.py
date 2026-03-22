"""
Servicio de importación del catálogo desde Excel.

PROBLEMA RESUELTO (v3):
- El Excel es un .xlsm con FÓRMULAS en columnas clave (PRECIO PUBLICO NETO,
  CODIGO FERROL, DESCRIPCION FERROL, todos los precios por descuento).
- pandas.read_excel lee fórmulas como strings ('=ROUND(M2,0)') a menos que
  se use engine='openpyxl' + data_only=True via openpyxl directamente.
- La hoja correcta es 'PRECIOS', no la primera hoja ('Hoja1' está vacía).

SOLUCIÓN:
- Leer con openpyxl load_workbook(data_only=True) para obtener valores calculados.
- Buscar automáticamente la hoja que contenga los encabezados esperados.
- Normalización robusta de encabezados.
"""

import logging
import re
import unicodedata
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.db.database import AsyncSessionLocal
from app.models.producto import Producto

logger = logging.getLogger(__name__)


# ── Mapeo de columnas normalizadas → campos del modelo ────────────────────────
MAPEO_COLUMNAS: dict[str, str] = {
    "producto": "producto",
    "descripcion": "descripcion",
    "descripción": "descripcion",
    "marca": "marca",
    "categoria": "categoria",
    "categoría": "categoria",
    "unidad": "unidad",
    "multiplo": "multiplo",
    "múltiplo": "multiplo",
    "master": "master",
    "precio lista": "precio_lista",
    "preciolista": "precio_lista",
    "promocion": "promocion",
    "promoción": "promocion",
    "vigencia": "vigencia",
    "costo + iva": "costo_mas_iva",
    "costo+iva": "costo_mas_iva",
    "costo mas iva": "costo_mas_iva",
    "iva": "iva",
    "precio sugerido de venta con iva": "precio_sugerido_con_iva",
    "precio sugerido con iva": "precio_sugerido_con_iva",
    "codigo sat": "codigo_sat",
    "código sat": "codigo_sat",
    # Con espacio al final (como viene en el Excel real: 'PRECIO PUBLICO NETO ')
    "precio publico neto": "precio_publico_neto",
    "precio público neto": "precio_publico_neto",
    "codigo ferrol": "codigo_ferrol",
    "código ferrol": "codigo_ferrol",
    "espacio": "espacio",
    "descripcion ferrol": "descripcion_ferrol",
    "descripción ferrol": "descripcion_ferrol",
    "precio 20%": "precio_20",
    "precio 8.5%": "precio_85",
    "precio 12%": "precio_12",
    "precio publico": "precio_publico",
    "precio público": "precio_publico",
    "precio publico con 5% descuento": "precio_publico_5_desc",
    "precio público con 5% descuento": "precio_publico_5_desc",
}

CAMPOS_NUMERICOS = {
    "precio_lista", "promocion", "costo_mas_iva", "iva",
    "precio_sugerido_con_iva", "precio_publico_neto",
    "precio_20", "precio_85", "precio_12",
    "precio_publico", "precio_publico_5_desc",
}


def _normalizar_encabezado(texto) -> str:
    """Normaliza un encabezado: minúsculas, sin tildes, sin espacios extra."""
    if texto is None:
        return ""
    texto = str(texto).strip()
    texto = re.sub(r"[\n\r\t]+", " ", texto)
    texto = re.sub(r"\s+", " ", texto).lower()
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in nfkd if not unicodedata.combining(c)).strip()


def _convertir_numero(valor) -> Optional[float]:
    """Convierte un valor a float de forma robusta."""
    if valor is None:
        return None
    import math
    if isinstance(valor, float) and math.isnan(valor):
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    if isinstance(valor, str):
        limpio = valor.strip().replace("$", "").replace(",", "").replace(" ", "")
        if limpio.startswith("(") and limpio.endswith(")"):
            limpio = "-" + limpio[1:-1]
        # Si empieza con '=' es una fórmula no calculada — ignorar
        if limpio.startswith("="):
            return None
        try:
            return float(limpio)
        except ValueError:
            return None
    return None


def _normalizar_texto_busqueda(producto: Producto) -> str:
    """Genera texto combinado normalizado para matching."""
    partes = []
    for campo in ["producto", "descripcion", "marca", "categoria",
                  "descripcion_ferrol", "codigo_ferrol"]:
        valor = getattr(producto, campo, None)
        if valor and isinstance(valor, str) and not valor.startswith("="):
            partes.append(valor.strip())
    return _normalizar_encabezado(" ".join(partes))


def _encontrar_hoja_correcta(wb) -> str:
    """
    Busca automáticamente la hoja que contenga los encabezados del catálogo.
    Prioriza la hoja configurada en EXCEL_HOJA; si no, busca la que tenga
    'Producto' o 'Descripción' en la primera fila.
    """
    # Si el usuario configuró una hoja específica, usarla
    if settings.EXCEL_HOJA and settings.EXCEL_HOJA in wb.sheetnames:
        logger.info(f"Usando hoja configurada: '{settings.EXCEL_HOJA}'")
        return settings.EXCEL_HOJA

    # Buscar automáticamente la hoja con encabezados de catálogo
    palabras_clave = {"producto", "descripcion", "descripción", "marca", "precio"}
    for nombre in wb.sheetnames:
        ws = wb[nombre]
        primera_fila = next(ws.iter_rows(max_row=1, values_only=True), None)
        if primera_fila:
            encabezados = {_normalizar_encabezado(v) for v in primera_fila if v}
            coincidencias = encabezados & palabras_clave
            if len(coincidencias) >= 2:
                logger.info(f"Hoja detectada automáticamente: '{nombre}' (coincidencias: {coincidencias})")
                return nombre

    # Fallback: primera hoja
    logger.warning(f"No se encontró hoja con encabezados conocidos. Usando primera: '{wb.sheetnames[0]}'")
    return wb.sheetnames[0]


class CatalogoService:

    async def cargar_catalogo_desde_excel(self) -> int:
        ruta = Path(settings.EXCEL_RUTA)
        if not ruta.exists():
            raise FileNotFoundError(
                f"Excel no encontrado: {ruta.absolute()}\n"
                f"Configura EXCEL_RUTA en el archivo .env"
            )

        logger.info(f"Leyendo Excel: {ruta.absolute()}")

        # ── CLAVE: data_only=True para leer valores de fórmulas, no las fórmulas ──
        # Sin esto, columnas como PRECIO PUBLICO NETO y CODIGO FERROL
        # vendrían como '=ROUND(M2,0)' y '=CONCATENATE(...)' respectivamente.
        try:
            wb = load_workbook(str(ruta), read_only=True, data_only=True)
        except Exception as e:
            raise Exception(f"Error al abrir Excel '{ruta}': {e}")

        nombre_hoja = _encontrar_hoja_correcta(wb)
        ws = wb[nombre_hoja]

        # Leer encabezados de la fila configurada
        fila_enc = settings.EXCEL_FILA_ENCABEZADO + 1  # openpyxl es 1-indexed
        encabezados_raw = []
        for row in ws.iter_rows(min_row=fila_enc, max_row=fila_enc, values_only=True):
            encabezados_raw = list(row)
            break

        logger.info(f"Encabezados encontrados ({len(encabezados_raw)}): {encabezados_raw[:10]}...")

        # Construir mapeo índice → campo modelo
        mapeo_indices: dict[int, str] = {}
        for idx, enc in enumerate(encabezados_raw):
            enc_norm = _normalizar_encabezado(enc)
            if enc_norm in MAPEO_COLUMNAS:
                campo = MAPEO_COLUMNAS[enc_norm]
                mapeo_indices[idx] = campo
                logger.debug(f"  ✓ Col {idx+1} '{enc}' → {campo}")
            elif enc is not None:
                logger.warning(f"  ✗ Col {idx+1} sin mapeo: '{enc}' (normalizada: '{enc_norm}')")

        logger.info(f"Columnas mapeadas: {len(mapeo_indices)} de {len(encabezados_raw)}")

        # Cargar productos
        cantidad = await self._guardar_productos_openpyxl(ws, mapeo_indices, fila_enc)
        wb.close()
        return cantidad

    async def _guardar_productos_openpyxl(
        self,
        ws,
        mapeo_indices: dict[int, str],
        fila_encabezado: int,
    ) -> int:
        async with AsyncSessionLocal() as db:
            await db.execute(text("DELETE FROM productos"))
            await db.commit()
            logger.info("Tabla limpiada. Iniciando carga...")

            cargados = 0
            omitidos = 0
            fila_num = fila_encabezado  # contador para referencia

            for row in ws.iter_rows(min_row=fila_encabezado + 1, values_only=True):
                fila_num += 1

                # Omitir filas vacías
                if all(v is None for v in row):
                    continue

                p = Producto()
                p.fila_excel = fila_num

                for idx, campo in mapeo_indices.items():
                    if idx >= len(row):
                        continue
                    valor = row[idx]

                    if campo in CAMPOS_NUMERICOS:
                        setattr(p, campo, _convertir_numero(valor))
                    else:
                        # Campo de texto — ignorar fórmulas no calculadas
                        if valor is None:
                            setattr(p, campo, None)
                        elif isinstance(valor, str) and valor.startswith("="):
                            setattr(p, campo, None)
                        elif isinstance(valor, str) and valor.strip() in ("", "nan", "None"):
                            setattr(p, campo, None)
                        else:
                            setattr(p, campo, str(valor).strip())

                # Requiere al menos nombre o descripción
                if not p.producto and not p.descripcion:
                    omitidos += 1
                    continue

                p.texto_busqueda = _normalizar_texto_busqueda(p)
                db.add(p)
                cargados += 1

                if cargados % 500 == 0:
                    await db.commit()
                    logger.info(f"  ... {cargados} productos cargados")

            await db.commit()
            logger.info(f"Carga completa: {cargados} productos, {omitidos} omitidos.")
            return cargados

    async def obtener_total_productos(self) -> int:
        from sqlalchemy import select, func
        async with AsyncSessionLocal() as db:
            r = await db.execute(select(func.count(Producto.id)))
            return r.scalar() or 0
